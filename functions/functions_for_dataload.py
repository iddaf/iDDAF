##########################################################################################
def generate_others(numOfBuses, numOfLines, num_Of_attacked_buses, data, Topo):
    import panda as pd
    import numpy as np
    data = pd.DataFrame(data)
    numOfZ = numOfBuses + 2* numOfLines
    num_Of_attack = int (data.shape[0]/numOfZ)
    # Preprocessing data

    false_Data = pd.DataFrame([])
    combination_Data_Sensor = pd.DataFrame([])
    combination_Data_Cluster = pd.DataFrame([])

    start_indx = 0
    end_indx = start_indx + numOfZ + 1

    progress = tqdm.tqdm(total=num_Of_attack, desc='Done', position=0)
    while (end_indx <= data.shape[0]):
        progress.update(1)
        attack_data = data.iloc[start_indx + 1 : end_indx, 2]
        combination_data = data.iloc[start_indx + 1 : end_indx, 1].astype(int)

        sensor_ids = np.where(combination_data.values == 1)[0]+1
        if len(sensor_ids) > 0:        
            cluster_maping,_ =  assignClusterID (sensor_ids, Topo, NumOfBuses=numOfBuses, NumOfLines = numOfLines)
            combination_Data_Cluster = combination_Data_Cluster.append(pd.DataFrame(cluster_maping, index = None, columns = None).T)

            false_Data = false_Data.append(pd.DataFrame(attack_data.values, index = None, columns = None).T)
            combination_Data_Sensor = combination_Data_Sensor.append(pd.DataFrame(combination_data.values, index = None, columns = None).T)

        start_indx += numOfZ+1
        end_indx = start_indx + numOfZ + 1
        #print(start_indx, end_indx)
    false_Data.index = range(1, false_Data.shape[0]+1)
    false_Data.columns = range(1, false_Data.shape[1]+1)
    combination_Data_Sensor.index = range(1, combination_Data_Sensor.shape[0]+1)
    combination_Data_Sensor.columns = range(1, combination_Data_Sensor.shape[1]+1)

    combination_Data_Cluster.index = range(1, combination_Data_Cluster.shape[0]+1)
    combination_Data_Cluster.columns = range(1, combination_Data_Cluster.shape[1]+1)


    false_Data.to_excel(f'Attack Data//false_Data_{numOfBuses}_{numOfLines}_{num_Of_attacked_buses}.xlsx', 
                        engine='xlsxwriter', index=True) 
    combination_Data_Sensor.to_excel(f'Attack Data//combination_Data_Sensor_{numOfBuses}_{numOfLines}_{num_Of_attacked_buses}.xlsx',
                                     engine='xlsxwriter', index = True) 
    combination_Data_Cluster.to_excel(f'Attack Data//combination_Data_Cluster_{numOfBuses}_{numOfLines}_{num_Of_attacked_buses}.xlsx',
                                      engine='xlsxwriter', index = True) 
    return false_Data, combination_Data_Sensor, combination_Data_Cluster

##########################################################################################
def preprocess_data(bus_data_df, line_data_df):
    import pandas as pd
    import numpy as np
    import openpyxl
    from openpyxl import load_workbook

    # update the Remote controlled bus number with index
    bus_data_df['Remote controlled bus number'] = bus_data_df.index.values

    for index, bus_ in enumerate(line_data_df[['Tap bus number', 'Z bus number']].values):

        # upadate the Tap bus number
        try:
            location_0 = pd.Index(bus_data_df['Bus number']).get_loc(bus_[0])
            location_1 = pd.Index(bus_data_df['Bus number']).get_loc(bus_[1])

        except:
            print("Error! Index was not found!")
            break

        # following the pattern
        if location_0 <= location_1:
            fromBus = location_0
            toBus = location_1
        else:
            fromBus = location_1
            toBus = location_0

        line_data_df['Tap bus number'].iloc[index] = bus_data_df['Remote controlled bus number'].iloc[fromBus].copy()
        line_data_df['Z bus number'].iloc[index] = bus_data_df['Remote controlled bus number'].iloc[toBus].copy()

        # sort based on Tab bus number and Z bus number
    line_data_df.sort_values(by=['Tap bus number', 'Z bus number'], inplace=True)


#####################################################################################
def generate_topology_matrix(numOfBuses, numOfLines,line_data_df, file_name):
    import pandas as pd
    import numpy as np
    import openpyxl
    from openpyxl import load_workbook
    
    numOfZ = numOfBuses + numOfLines*2

    # Create line data and topology matrix
    line_data = line_data_df[['Tap bus number', 'Z bus number', 'X']]

    # calculate admittance of lines
    try:
        line_data['B'] = round(1 / line_data['X'], 2)
    except:
        print("Got 0 as reactance! Inversion is not possible")

    # removing X from the data frame as we wont need it.
    line_data.insert(loc=0, column='Line ID', value= np.arange(1, numOfLines+1))

    line_data = line_data.drop(columns=['X'], axis=1)

    # Calling topology processor to calcuate topology matrix
    topo_mat = topologyProcessor(numOfBuses, numOfLines, line_data)

    # Saving the Line data and Topology
    book = load_workbook(file_name)
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    writer.book = book
    line_data.to_excel(writer, 'Line Data', index=False)
    topo_mat.to_excel(writer, 'Topology Matrix', index=False)
    writer.save()
    writer.close()
    # saving complete !

    return topo_mat, line_data


##################################################################
# Topolgy matrix generator --> takes line data return topomat
def topologyProcessor(numOfBuses, numOfLines, line_data):
    import pandas as pd
    import numpy as np
    import openpyxl
    from openpyxl import load_workbook

    numOfZ = numOfBuses + numOfLines * 2

    # Placeholder for topoMat
    topo_mat = pd.DataFrame(np.zeros((numOfZ + 1, numOfBuses), dtype=float))

    # rows representing the line powerflow
    for line in line_data.values:
        # Line information
        lineID = int(line[0])
        fromBus = int(line[1] - 1)
        toBus = int(line[2] - 1)
        admittance = line[3]

        #
        topo_mat.iloc[lineID, fromBus] = admittance
        topo_mat.iloc[lineID, toBus] = - admittance

        topo_mat.iloc[lineID + numOfLines, fromBus] = - admittance
        topo_mat.iloc[lineID + numOfLines, toBus] = admittance

        # rows representing the bus consumption
    for busIndx in range(1, numOfBuses + 1):

        busTopo = np.zeros(numOfBuses)

        for line in line_data.values:

            # Line information
            lineID = int(line[0])
            fromBus = int(line[1])
            toBus = int(line[2])

            if fromBus == busIndx:
                busTopo = busTopo + topo_mat.loc[lineID]

            elif toBus == busIndx:
                busTopo = busTopo - topo_mat.loc[lineID]

        topo_mat.loc[2 * numOfLines + busIndx] = busTopo.copy()

        # adding 1 in the first line which represents the reference bus
    topo_mat.iloc[0, 0] = 1
    return topo_mat


# Generate Z_msr and Bus_data for ideal case
def generate_Z_msr_org(numOfBuses, numOfLines, bus_data_df, topo_mat, file_name):
    import pandas as pd
    import numpy as np
    import openpyxl
    from openpyxl import load_workbook
    
    # Creating Measurement Data to run state estimation
    bus_data = bus_data_df[['Remote controlled bus number', 'Load MW', 'Generation MW']]
    bus_data.columns = ['Bus number', 'Load', 'Generation']

    # Correcting the load generation for a lossless DC system
    correction_load = sum(bus_data['Load']) - sum(bus_data['Generation'])
    print("correction_load: ", correction_load)

    # Adding the correction load to the largest generator
    bus_data['Generation'].loc[bus_data['Generation'].idxmax()] += correction_load
    # correction_check = sum(bus_data['Load']) - sum(bus_data['Generation'])
    # print("correction_check: ", correction_check)

    # Bus Power = Bus Gen - Bus Load
    bus_data['Bus Power'] = bus_data['Generation'] - bus_data['Load']

    print("bus_data:\n", bus_data.head())

    # Padding 0 in the top of the data from reference
    Z_data_bus_power = pd.DataFrame(pd.concat([pd.Series([0]), bus_data['Bus Power']]))

    # Topomat containing only the bus power rows along with reference bus
    B_mat_bus_power = pd.concat([topo_mat.loc[0:0], topo_mat.loc[numOfLines * 2 + 1:]])

    # Estimating the states fromt the bus power data
    state_original = np.linalg.pinv(B_mat_bus_power) @ Z_data_bus_power

    # Calculating the Z_msr_org using the Topology Matrix and states
    Z_msr_org = topo_mat @ state_original
    Z_msr_org.columns = ['Data']
    # Saving the data
    book = load_workbook(file_name)
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    writer.book = book
    Z_msr_org.to_excel(writer, "Measurement Data", index=False)
    bus_data.to_excel(writer, "Bus Data", index=False)
    writer.save()
    writer.close()
    # saving complete !

    print("Z_msr_org:\n", Z_msr_org.head())

    return Z_msr_org, bus_data
##############################################


def Convert2Float(list_):
    floatList = []
    
    for elements in list_.strip().split(" "):
        elementParts = elements.split("/")
        if len(elementParts) > 1:
            value = float (elementParts[0]) / float (elementParts[1])
        else:
            value = float (elementParts[0])
        floatList.append(value)
        
    return floatList
##########################################################3
def checkAccess(AttackMat, Access):
    import numpy as np
    return np.where((Access[np.where(AttackMat[:,1]==1)]==1).flatten() == False)[0].size == 0
##########################################################################

def generate_attackdata(numOfBuses, numOfLines, line_data, attacked_Bus, current_path):

    import os
    import numpy as np
    import psutil    
    import glob

    ##########################################################################3
#     numOfBuses = 14
#     numOfLines = 20
    numOfZ = numOfBuses + numOfLines*2
#     attacked_Bus = 7

    ## Creating the input file for attack generation

    # File name and directory
    file = open("Attack Gen\\Input_Topo.txt", 'w+')

    # Printing numOfBuses, numOfLines, Reference Bus 
    file.write("# numOfBuses, numOfLines, Reference Bus\n")
    line = f"{numOfBuses}\t{numOfLines}\t1\n\n"
    file.write(line)

    # Printing #Lines Info
    file.write("\n#Lines Info\n")

    # Line Data contains the line information
    data = line_data.values.copy()
    data = data.astype(str).tolist()
    for line in data:
        line += ['1']*4
        s = '\t'.join(line) 
        file.write(s)
        file.write("\n")

    # Printing Measurements info
    file.write("\n#Measurements Info\n")
    for i in range(1, numOfZ+1):
        string = f"{i}\t1\t0\t1\n"
        file.write(string)

    # #Number of States 
    file.write("\n\n#Number of States\n")
    file.write("0")

    # Maximum number of states for estimation attack
    file.write("\n\n#Maximum number of states\n")
    file.write(f"{numOfBuses-1}\n")
    file.write("0\n")

    ## Attacker's Resource Limitation
    file.write("\n#Attacker's Resource Limitation, Measurements, Bus\n")
    file.write(f"{attacked_Bus*5}\t{attacked_Bus}")
    file.close()
    print("Input_Topo.txt---> Generated!")
    print("attacked_Bus: ", attacked_Bus)
    # ###################################################################

    print("Starting.... Prog_State_Estimation.exe")
    # current_path = os.getcwd()
    attack_path = current_path + "//Attack Gen//"
    # path = r"C://Users//shahr//Dropbox//Shared with Shahriar//Deceiption State Estimation//Other Codes//IEEE Bus Data//Attack Gen//"
    os.chdir(attack_path)
    os.system("Prog_State_Estimation.exe")
    print("Attack Data Created.....\n ")
    os.chdir(current_path)
    # #######################################################################


    #########################################################
    mydir = "Attack Gen\\"
    file_list = glob.glob(mydir+"Attack_Bus_Vectors_%d_%d_%d.txt"%(numOfBuses, numOfLines, attacked_Bus))
    #file_list = glob.glob(mydir+"*.txt")
    print('file_list {}'.format(file_list))
    ############

    ################################################################
    for fileName in file_list:

        nameParts = fileName.split("_")
        # initiating the attack matrix 
        AttackMat = np.zeros ((numOfZ+1, 3), dtype = float)
        AttackMat[:,0] = np.arange(0,numOfZ+1) 
        #####################################

        file = open(fileName, "r")
        lines = file.readlines()

        ############################
        startFlag = 1
        doneFlag = 0
        saveIndx = 0
        statesFound = 0

        attackedStates = []
        for line in lines:

            if doneFlag == 1:
                doneFlag = 0
                if startFlag == 1: #and clear == True:
                    AttackSpace = AttackMat.copy() 
                    startFlag = 0
                    #print("Got the first attack vector! Attack space initiated!\n\n")
                else:
                    AttackSpace = np.vstack((AttackSpace, AttackMat))
                    #print("Updating attack space\n\n")

                AttackMat = np.zeros ((numOfZ+1, 3), dtype = float)
                AttackMat[:,0] = np.arange(0,numOfZ+1) 

            parts = line.split (":")

            try:
                if parts[0] == "The measurements that are required to alter":
                    #print("The measurements that are required to alter:")
                    saveIndx = 1

                elif saveIndx == 1:
                    if len(parts[0].strip()) > 0:
                        partMat = np.array(Convert2Float(parts[0])).astype(int)
                        AttackMat [partMat, 1] = 1
                        pass
                    else:
                        pass
                        #print("Nothing attacked")
                    saveIndx = 0

                if statesFound == 1:
                    #rint("Found States:", line)
                    attackedStates.append(line.strip())
                    statesFound = 0

                if parts[0] =="The states that are attacked":
                    statesFound = 1

                if parts[0] == "Forward line flows":
                    #print("Got Forward >>>", end = " ")
                    partMat = np.mat(Convert2Float(parts[1]))
                    AttackMat [1:numOfLines+1,2] = partMat

                elif parts[0] == "Backward line flows":
                    #print("Got Backward >>>", end = " ")

                    partMat = np.mat(Convert2Float(parts[1]))
                    AttackMat [numOfLines+1 : 2*numOfLines+1, 2] = partMat

                elif parts[0] == "Bus power consumptions":

                    #print("Got Bus power !!")
                    partMat = np.mat(Convert2Float(parts[1]))
                    AttackMat [ 2*numOfLines+1 :, 2] = partMat
                    AttackMat[np.where(AttackMat[:,1] !=1), 2] = 0
                    doneFlag = 1
            except:
                print("************Incomplete Attack Vector**************")

            ####################################################

        try:
            outputfname = current_path+f"\\Attack Data\\Attack_Space_{numOfBuses}_{numOfLines}_{attacked_Bus}.csv"
            print("Saved File: ",outputfname)
            np.savetxt(outputfname, AttackSpace, fmt = "%f", delimiter =',')
            return AttackSpace
        except:
            print("Error Saving File!")
            return None
